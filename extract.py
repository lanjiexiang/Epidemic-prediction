import os
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook

# 设置文件的绝对路径
file_path = r'C:\\CodeSandbox\\Shandong2023\\山东2023年度无害化数据.xlsx'

# 检查文件是否存在
if not os.path.exists(file_path):
    print(f"文件 {file_path} 不存在，请检查路径是否正确")
else:
    print("开始读取数据")

    # 打开Excel文件
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.active

    # 获取总行数（包含标题行）
    total_rows = sheet.max_row

    # 使用 tqdm 显示读取进度
    data = []
    with tqdm(total=total_rows-1, desc="Reading Excel File") as pbar:
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过标题行
            data.append(row)
            pbar.update(1)

    # 将数据转换为DataFrame
    df = pd.DataFrame(data, columns=[cell.value for cell in sheet[1]])

    print("数据读取完成")

    # 提取畜种（第2列）、动物死亡数量（第3列）、死亡时间（第6列）和养殖场户（第5列）
    animal_types = df.iloc[:, 1]
    death_counts = df.iloc[:, 2]
    farm = df.iloc[:, 4]
    death_times = df.iloc[:, 5]

    # 将死亡时间列转换为日期格式，只保留日期部分
    death_dates = pd.to_datetime(death_times, errors='coerce').dt.date

    # 创建DataFrame
    data = pd.DataFrame({
        'animal_type': animal_types, 
        'death_date': death_dates, 
        'death_count': death_counts, 
        'farm': farm
    })

    # 处理缺失日期（dropna函数）
    data = data.dropna(subset=['death_date'])

    # 只保留牛的数据
    cow_data = data[data['animal_type'] == '牛']

    # 按死亡日期和养殖场分组并求和
    result = cow_data.groupby(['death_date', 'farm']).sum().reset_index()

    # 取前100个屠宰场
    top_100_farms = result['farm'].value_counts().nlargest(100).index
    result_top_100 = result[result['farm'].isin(top_100_farms)]

    # 创建透视表，横向索引是屠宰场，纵向索引是死亡日期
    pivot_table = result_top_100.pivot_table(
        index='death_date', 
        columns='farm', 
        values='death_count', 
        aggfunc='sum',
        fill_value=0
    )

    # 输出透视表
    print(pivot_table)

    # 将透视表保存到新的Excel文件
    output_file_path = r'C:\\CodeSandbox\\Shandong2023\\牛死亡数量统计.xlsx'
    pivot_table.to_excel(output_file_path)
    print(f"结果已保存到 {output_file_path}")
    
    # 关闭工作簿
    workbook.close()
