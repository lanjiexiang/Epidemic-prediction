import pandas as pd
import os
from tqdm import tqdm
from openpyxl import load_workbook

def extract_rows_by_date(input_file, date_to_extract, column_index=10):
    # 检查文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 文件未找到 - {input_file}")
        return

    # 使用 openpyxl 加载工作簿并获取总行数
    workbook = load_workbook(filename=input_file, read_only=True)
    sheet = workbook.active
    total_rows = sheet.max_row

    # 初始化进度条
    progress_bar = tqdm(total=total_rows, desc="读取文件", unit="行")

    # 用于存储符合条件的数据
    filtered_rows = []

    try:
        # 逐行读取数据
        for row in sheet.iter_rows(min_row=2, max_row=total_rows, values_only=True):
            progress_bar.update(1)
            # 转换日期列格式
            date_value = pd.to_datetime(row[column_index], errors='coerce').date()
            if date_value == pd.to_datetime(date_to_extract).date():
                filtered_rows.append(row)

        # 如果有过滤后的行，将其转换为 DataFrame
        if filtered_rows:
            filtered_df = pd.DataFrame(filtered_rows, columns=[cell.value for cell in sheet[1]])
        else:
            filtered_df = pd.DataFrame()

    except IndexError:
        print(f"错误: 列索引 '{column_index}' 超出范围。请确认列索引正确。")
        return
    except Exception as e:
        print(f"处理数据时出错: {e}")
        return
    finally:
        # 关闭进度条
        progress_bar.close()

    # 创建保存文件的目录
    output_folder = os.path.join(os.path.dirname(input_file), 'extracted_data')
    os.makedirs(output_folder, exist_ok=True)

    # 构造输出文件路径
    output_file = os.path.join(output_folder, f'extracted_{date_to_extract}.xlsx')

    try:
        # 将过滤后的数据保存到新的Excel文件
        print("保存提取数据...")
        filtered_df.to_excel(output_file, index=False)
        print(f'提取的数据已保存到 {output_file}')
    except Exception as e:
        print(f"保存数据时出错: {e}")

# 使用示例
input_file = r'C:\CodeSandbox\Shandong2023\山东2023年1-12月猪牛羊产地检疫信息.xlsx'  # 输入文件路径
date_to_extract = '2023-01-08'  # 要提取的日期
extract_rows_by_date(input_file, date_to_extract)